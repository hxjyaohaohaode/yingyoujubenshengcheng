"""
大纲架构生成与修改服务

支持:
- AI生成包含主线/支线的大纲架构图
- 自然语言修改大纲
- 大纲节点CRUD
- 大纲→章节/场景同步
"""
import json
import logging
from dataclasses import dataclass, field
from typing import Optional
from uuid import uuid4

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, text

from core.gateway.client import get_gateway
from config import DATABASE_URL

_IS_SQLITE = DATABASE_URL.startswith("sqlite")

def _now_expr() -> str:
    return "datetime('now')" if _IS_SQLITE else "NOW()"

logger = logging.getLogger(__name__)


@dataclass
class OutlineNode:
    id: str = ""
    node_type: str = "chapter"
    title: str = ""
    summary: str = ""
    position_x: float = 0
    position_y: float = 0
    parent_id: Optional[str] = None
    arc_type: str = "main"
    emotion_target: int = 5
    word_target: int = 0
    metadata: dict = field(default_factory=dict)


@dataclass
class OutlineEdge:
    id: str = ""
    source_id: str = ""
    target_id: str = ""
    edge_type: str = "sequence"
    label: str = ""
    metadata: dict = field(default_factory=dict)


@dataclass
class OutlineGraph:
    nodes: list[OutlineNode] = field(default_factory=list)
    edges: list[OutlineEdge] = field(default_factory=list)


async def ai_generate_outline(
    db: AsyncSession,
    project_id: str,
    genre: str = "",
    theme: str = "",
    core_contradiction: str = "",
    target_chapters: int = 10,
    narrative_structure: str = "three_act",
    user_description: str = "",
) -> OutlineGraph:
    """AI生成大纲架构图"""
    gateway = get_gateway()
    if not gateway:
        return OutlineGraph()

    from core.narrative.memory_loader import build_narrative_context
    narrative_context = await build_narrative_context(db, project_id)

    structure_templates = {
        "three_act": "三幕式结构（建置→对抗→解决）",
        "hero_journey": "英雄之旅（12阶段）",
        "save_cat": "救猫咪（15个节拍）",
        "hook_reversal": "钩子-反转螺旋（短剧专用）",
        "escalation": "爽点递进（网文/短剧）",
        "five_act": "五幕式结构（开端→上升→高潮→回落→结局）",
    }

    structure_desc = structure_templates.get(narrative_structure, narrative_structure)

    prompt = f"""你是一位专业编剧，专精{genre or '互动影游'}剧本的大纲架构设计。

{narrative_context}

【项目信息】
- 题材: {genre}
- 主题: {theme}
- 核心矛盾: {core_contradiction}
- 目标章节数: {target_chapters}
- 叙事结构: {structure_desc}
{"- 用户描述: " + user_description if user_description else ""}

请生成一个包含主线和支线的大纲架构图，输出JSON格式:

{{
  "nodes": [
    {{
      "id": "arc_1",
      "node_type": "story_arc",
      "title": "主线：XXX",
      "summary": "这条线讲述...",
      "arc_type": "main",
      "emotion_target": 7,
      "word_target": 300000,
      "metadata": {{}}
    }},
    {{
      "id": "arc_2",
      "node_type": "story_arc",
      "title": "支线：XXX",
      "summary": "这条线讲述...",
      "arc_type": "sub",
      "emotion_target": 5,
      "word_target": 100000,
      "metadata": {{}}
    }},
    {{
      "id": "ch_1",
      "node_type": "chapter",
      "title": "第1章：XXX",
      "summary": "本章概要...",
      "parent_id": "arc_1",
      "arc_type": "main",
      "emotion_target": 6,
      "word_target": 30000,
      "metadata": {{"key_turning_points": ["转折1"], "foreshadow_tasks": ["伏笔1"]}}
    }},
    {{
      "id": "evt_1",
      "node_type": "event",
      "title": "关键事件：XXX",
      "summary": "事件描述...",
      "parent_id": "ch_1",
      "arc_type": "main",
      "emotion_target": 8,
      "metadata": {{"event_type": "turning_point"}}
    }},
    {{
      "id": "choice_1",
      "node_type": "choice",
      "title": "抉择：XXX",
      "summary": "玩家面临的选择...",
      "parent_id": "evt_1",
      "arc_type": "main",
      "emotion_target": 7,
      "metadata": {{"options": ["选项A", "选项B"]}}
    }}
  ],
  "edges": [
    {{
      "id": "e1",
      "source_id": "arc_1",
      "target_id": "ch_1",
      "edge_type": "contains",
      "label": ""
    }},
    {{
      "id": "e2",
      "source_id": "ch_1",
      "target_id": "ch_2",
      "edge_type": "sequence",
      "label": "推进"
    }},
    {{
      "id": "e3",
      "source_id": "ch_1",
      "target_id": "evt_1",
      "edge_type": "contains",
      "label": ""
    }},
    {{
      "id": "e4",
      "source_id": "evt_1",
      "target_id": "choice_1",
      "edge_type": "leads_to",
      "label": "触发"
    }},
    {{
      "id": "e5",
      "source_id": "arc_1",
      "target_id": "arc_2",
      "edge_type": "crosses",
      "label": "交汇"
    }}
  ]
}}

要求:
1. 至少包含1条主线和1-2条支线
2. 主线至少{target_chapters}个章节节点
3. 每章至少1个关键事件节点
4. 在关键转折处设置选择节点
5. 主线和支线之间有交汇边
6. 每个节点有明确的情感目标和字数目标
7. 边类型: contains(包含), sequence(顺序), leads_to(导致), crosses(交汇), reverses(逆转)
"""

    try:
        response = await gateway.invoke(
            intent="write.outline",
            messages=[{"role": "user", "content": prompt}],
            cost_profile="quality",
            max_tokens=32000,
            temperature=0.7,
            use_cache=False,
        )
        return _parse_outline_response(response.content)
    except Exception as e:
        logger.error("AI大纲生成失败: %s", e)
        return OutlineGraph()


async def ai_modify_outline(
    db: AsyncSession,
    project_id: str,
    current_graph: OutlineGraph,
    instruction: str,
) -> OutlineGraph:
    """用自然语言修改大纲"""
    gateway = get_gateway()
    if not gateway:
        return current_graph

    nodes_json = json.dumps([
        {"id": n.id, "type": n.node_type, "title": n.title, "summary": n.summary,
         "parent": n.parent_id, "arc": n.arc_type, "emotion": n.emotion_target}
        for n in current_graph.nodes
    ], ensure_ascii=False, indent=2)

    edges_json = json.dumps([
        {"id": e.id, "source": e.source_id, "target": e.target_id,
         "type": e.edge_type, "label": e.label}
        for e in current_graph.edges
    ], ensure_ascii=False, indent=2)

    prompt = f"""你是一位专业编剧，现在需要根据用户的修改指令调整大纲架构图。

【当前大纲节点】
{nodes_json}

【当前大纲连线】
{edges_json}

【用户修改指令】
{instruction}

请输出修改后的完整大纲架构图JSON（格式与之前相同，包含nodes和edges数组）。
只输出JSON，不要解释。"""

    try:
        response = await gateway.invoke(
            intent="write.outline",
            messages=[{"role": "user", "content": prompt}],
            cost_profile="quality",
            max_tokens=32000,
            temperature=0.5,
            use_cache=False,
        )
        result = _parse_outline_response(response.content)
        if result.nodes:
            return result
        return current_graph
    except Exception as e:
        logger.error("AI大纲修改失败: %s", e)
        return current_graph


async def save_outline_graph(
    db: AsyncSession,
    project_id: str,
    graph: OutlineGraph,
) -> dict:
    """保存大纲架构图到数据库"""
    for node in graph.nodes:
        if not node.id:
            node.id = str(uuid4())

    for edge in graph.edges:
        if not edge.id:
            edge.id = str(uuid4())

    nodes_json = json.dumps([
        {
            "id": n.id, "node_type": n.node_type, "title": n.title,
            "summary": n.summary, "position_x": n.position_x, "position_y": n.position_y,
            "parent_id": n.parent_id, "arc_type": n.arc_type,
            "emotion_target": n.emotion_target, "word_target": n.word_target,
            "metadata": n.metadata,
        }
        for n in graph.nodes
    ], ensure_ascii=False)

    edges_json = json.dumps([
        {
            "id": e.id, "source_id": e.source_id, "target_id": e.target_id,
            "edge_type": e.edge_type, "label": e.label, "metadata": e.metadata,
        }
        for e in graph.edges
    ], ensure_ascii=False)

    try:
        await db.execute(
            text(f"""
                INSERT INTO project_outline_graphs (project_id, nodes_json, edges_json, updated_at)
                VALUES (:pid, :nodes, :edges, {_now_expr()})
                ON CONFLICT (project_id) DO UPDATE SET
                    nodes_json = :nodes, edges_json = :edges, updated_at = {_now_expr()}
            """),
            {"pid": project_id, "nodes": nodes_json, "edges": edges_json},
        )
        await db.commit()
    except Exception as e:
        logger.error("大纲保存失败: %s", e)
        try:
            await db.execute(
                text("""
                    CREATE TABLE IF NOT EXISTS project_outline_graphs (
                        project_id TEXT PRIMARY KEY,
                        nodes_json TEXT NOT NULL,
                        edges_json TEXT NOT NULL,
                        updated_at TEXT
                    )
                """)
            )
            await db.commit()
            await db.execute(
                text(f"""
                    INSERT INTO project_outline_graphs (project_id, nodes_json, edges_json, updated_at)
                    VALUES (:pid, :nodes, :edges, {_now_expr()})
                    ON CONFLICT (project_id) DO UPDATE SET
                        nodes_json = :nodes, edges_json = :edges, updated_at = {_now_expr()}
                """),
                {"pid": project_id, "nodes": nodes_json, "edges": edges_json},
            )
            await db.commit()
        except Exception as e2:
            logger.error("大纲保存重试失败: %s", e2)

    return {
        "project_id": project_id,
        "node_count": len(graph.nodes),
        "edge_count": len(graph.edges),
    }


async def load_outline_graph(
    db: AsyncSession,
    project_id: str,
) -> OutlineGraph:
    """从数据库加载大纲架构图"""
    try:
        result = await db.execute(
            text("SELECT nodes_json, edges_json FROM project_outline_graphs WHERE project_id = :pid"),
            {"pid": project_id},
        )
        row = result.fetchone()
        if not row:
            return OutlineGraph()

        nodes_data = json.loads(row[0]) if row[0] else []
        edges_data = json.loads(row[1]) if row[1] else []

        nodes = [
            OutlineNode(
                id=n.get("id", ""), node_type=n.get("node_type", "chapter"),
                title=n.get("title", ""), summary=n.get("summary", ""),
                position_x=n.get("position_x", 0), position_y=n.get("position_y", 0),
                parent_id=n.get("parent_id"), arc_type=n.get("arc_type", "main"),
                emotion_target=n.get("emotion_target", 5),
                word_target=n.get("word_target", 0),
                metadata=n.get("metadata", {}),
            )
            for n in nodes_data
        ]

        edges = [
            OutlineEdge(
                id=e.get("id", ""), source_id=e.get("source_id", ""),
                target_id=e.get("target_id", ""), edge_type=e.get("edge_type", "sequence"),
                label=e.get("label", ""), metadata=e.get("metadata", {}),
            )
            for e in edges_data
        ]

        return OutlineGraph(nodes=nodes, edges=edges)
    except Exception as e:
        logger.error("加载大纲架构图失败: %s", e)
        return OutlineGraph()


async def parse_document_to_outline(
    db: AsyncSession,
    project_id: str,
    file_content: bytes,
    filename: str,
) -> OutlineGraph:
    """解析上传的文档/图片，AI提取大纲架构"""
    gateway = get_gateway()
    if not gateway:
        return OutlineGraph()

    import base64
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "txt"

    text_content = ""
    if ext in ("png", "jpg", "jpeg"):
        b64 = base64.b64encode(file_content).decode()
        mime = f"image/{'jpeg' if ext == 'jpg' else ext}"
        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": _PARSE_OUTLINE_PROMPT},
                    {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
                ],
            }
        ]
    else:
        if ext == "pdf":
            text_content = _extract_pdf_text(file_content)
        elif ext in ("doc", "docx"):
            text_content = _extract_docx_text(file_content)
        elif ext in ("xls", "xlsx"):
            text_content = _extract_xlsx_text(file_content)
        else:
            text_content = file_content.decode("utf-8", errors="ignore")

        if not text_content.strip():
            return OutlineGraph()

        messages = [{"role": "user", "content": f"{_PARSE_OUTLINE_PROMPT}\n\n【文档内容】\n{text_content}"}]

    try:
        response = await gateway.invoke(
            intent="write.outline",
            messages=messages,
            cost_profile="quality",
            max_tokens=32000,
            temperature=0.3,
            use_cache=False,
        )
        return _parse_outline_response(response.content)
    except Exception as e:
        logger.error("文档解析大纲失败: %s", e)
        return OutlineGraph()


_PARSE_OUTLINE_PROMPT = """你是一位专业编剧助手，擅长从文档中提取大纲架构。

请仔细分析提供的文档内容（可能是图片或文本），提取其中的大纲结构，输出JSON格式:

{
  "nodes": [
    {"id": "arc_1", "node_type": "story_arc", "title": "主线：XXX", "summary": "...", "arc_type": "main", "emotion_target": 7, "word_target": 300000, "metadata": {}},
    {"id": "ch_1", "node_type": "chapter", "title": "第1章：XXX", "summary": "...", "parent_id": "arc_1", "arc_type": "main", "emotion_target": 6, "word_target": 30000, "metadata": {}},
    {"id": "evt_1", "node_type": "event", "title": "关键事件：XXX", "summary": "...", "parent_id": "ch_1", "arc_type": "main", "emotion_target": 8, "metadata": {"event_type": "turning_point"}},
    {"id": "choice_1", "node_type": "choice", "title": "抉择：XXX", "summary": "...", "parent_id": "evt_1", "arc_type": "main", "emotion_target": 7, "metadata": {"options": ["选项A", "选项B"]}}
  ],
  "edges": [
    {"id": "e1", "source_id": "arc_1", "target_id": "ch_1", "edge_type": "contains", "label": ""},
    {"id": "e2", "source_id": "ch_1", "target_id": "ch_2", "edge_type": "sequence", "label": "推进"},
    {"id": "e3", "source_id": "ch_1", "target_id": "evt_1", "edge_type": "contains", "label": ""},
    {"id": "e4", "source_id": "evt_1", "target_id": "choice_1", "edge_type": "leads_to", "label": "触发"}
  ]
}

要求:
1. 识别文档中的故事线、章节、事件、抉择等层次
2. 保持文档原有的层级关系和逻辑顺序
3. 边类型: contains(包含), sequence(顺序), leads_to(导致), crosses(交汇), foreshadows(伏笔), triggers(触发), reverses(逆转), conflicts(冲突), echoes(呼应), depends(依赖)
4. 只输出JSON，不要解释"""


def _extract_pdf_text(content: bytes) -> str:
    try:
        import fitz
        doc = fitz.open(stream=content, filetype="pdf")
        return "\n".join(page.get_text() for page in doc)
    except ImportError:
        try:
            from PyPDF2 import PdfReader
            import io
            reader = PdfReader(io.BytesIO(content))
            return "\n".join(page.extract_text() or "" for page in reader.pages)
        except ImportError:
            logger.warning("PyPDF2未安装，PDF解析不可用")
            return ""


def _extract_docx_text(content: bytes) -> str:
    try:
        from docx import Document
        import io
        doc = Document(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        logger.warning("python-docx未安装，DOCX解析不可用")
        return ""


def _extract_xlsx_text(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
        import io
        wb = load_workbook(io.BytesIO(content))
        lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                lines.append(" | ".join(str(c) for c in row if c is not None))
        return "\n".join(lines)
    except ImportError:
        logger.warning("openpyxl未安装，Excel解析不可用")
        return ""
    except Exception as e:
        logger.warning("Excel解析失败: %s", e)
        return ""


async def sync_outline_to_chapters(
    db: AsyncSession,
    project_id: str,
    graph: OutlineGraph,
) -> dict:
    """将大纲架构图同步为章节/节结构"""
    chapter_nodes = [n for n in graph.nodes if n.node_type == "chapter"]
    chapter_nodes.sort(key=lambda n: n.title)

    synced = 0
    for idx, ch_node in enumerate(chapter_nodes):
        chapter_number = idx + 1

        existing = await db.execute(
            text("SELECT id FROM chapters WHERE project_id = :pid AND chapter_number = :num"),
            {"pid": project_id, "num": chapter_number},
        )
        row = existing.fetchone()

        if row:
            await db.execute(
                text("""UPDATE chapters SET
                    title = :title, summary = :summary, outline = :outline,
                    emotion_target = :emotion, core_conflict = :conflict
                WHERE id = :cid"""),
                {
                    "title": ch_node.title,
                    "summary": ch_node.summary,
                    "outline": ch_node.summary,
                    "emotion": ch_node.emotion_target,
                    "conflict": ch_node.metadata.get("core_conflict", ""),
                    "cid": row[0],
                },
            )
        else:
            await db.execute(
                text("""INSERT INTO chapters
                    (id, project_id, chapter_number, title, summary, outline, emotion_target, status)
                    VALUES (:id, :pid, :num, :title, :summary, :outline, :emotion, 'draft')"""),
                {
                    "id": str(uuid4()),
                    "pid": project_id,
                    "num": chapter_number,
                    "title": ch_node.title,
                    "summary": ch_node.summary,
                    "outline": ch_node.summary,
                    "emotion": ch_node.emotion_target,
                },
            )
        synced += 1

    await db.commit()
    return {"synced_chapters": synced}


def _parse_outline_response(content: str) -> OutlineGraph:
    """解析LLM返回的大纲JSON"""
    if not content:
        return OutlineGraph()

    text = content.strip()
    if text.startswith("```"):
        lines = text.split("\n")
        if lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        text = "\n".join(lines)

    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        import re
        match = re.search(r'\{.*\}', text, re.DOTALL)
        if match:
            try:
                data = json.loads(match.group())
            except json.JSONDecodeError:
                return OutlineGraph()
        else:
            return OutlineGraph()

    nodes = []
    for n in data.get("nodes", []):
        nodes.append(OutlineNode(
            id=n.get("id", str(uuid4())),
            node_type=n.get("node_type", "chapter"),
            title=n.get("title", ""),
            summary=n.get("summary", ""),
            position_x=n.get("position_x", 0),
            position_y=n.get("position_y", 0),
            parent_id=n.get("parent_id"),
            arc_type=n.get("arc_type", "main"),
            emotion_target=n.get("emotion_target", 5),
            word_target=n.get("word_target", 0),
            metadata=n.get("metadata", {}),
        ))

    edges = []
    for e in data.get("edges", []):
        edges.append(OutlineEdge(
            id=e.get("id", str(uuid4())),
            source_id=e.get("source_id", ""),
            target_id=e.get("target_id", ""),
            edge_type=e.get("edge_type", "sequence"),
            label=e.get("label", ""),
            metadata=e.get("metadata", {}),
        ))

    return OutlineGraph(nodes=nodes, edges=edges)